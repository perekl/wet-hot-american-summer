"""Persist screenplay project changes to cues.json and master_cues.xlsx."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"

CUE_COLS = [
    "Cue ID", "Asset ID", "Cue Type", "Expected Background Asset", "Paragraph ID",
    "Script Page", "Scene", "Trigger Dialogue", "Cue Name",
    "Category", "Priority", "Duration", "Loop", "Fade", "Volume", "Notes",
]


def _style_header(ws, cols):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", size=11)
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def _cue_to_row(cue: dict) -> dict:
    return {
        "Cue ID": cue.get("id", ""),
        "Asset ID": cue.get("asset_id", ""),
        "Cue Type": cue.get("cue_type", ""),
        "Expected Background Asset": cue.get("expected_background_asset_id", ""),
        "Paragraph ID": cue.get("paragraph_id", ""),
        "Script Page": cue.get("page", ""),
        "Scene": cue.get("scene", ""),
        "Trigger Dialogue": cue.get("trigger", ""),
        "Cue Name": cue.get("name", ""),
        "Category": cue.get("category", ""),
        "Priority": cue.get("priority", ""),
        "Duration": cue.get("duration", ""),
        "Loop": cue.get("loop", ""),
        "Fade": cue.get("fade", ""),
        "Volume": cue.get("volume", ""),
        "Notes": cue.get("notes", ""),
    }


def write_master_cues_xlsx(cues: list[dict], path: Path = DATA / "master_cues.xlsx") -> None:
    rows = [_cue_to_row(c) for c in cues]
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Cues"
    _style_header(ws, CUE_COLS)
    for r, row in enumerate(rows, 2):
        for c, col in enumerate(CUE_COLS, 1):
            cell = ws.cell(row=r, column=c, value=row.get(col, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, col in enumerate(CUE_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(len(col) + 2, 40))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def persist_project(project) -> None:
    """Write cues.json and master_cues.xlsx from a ScriptProject."""
    cues = project.get_cues_sorted()
    project.cues_path.write_text(json.dumps(cues, indent=2), encoding="utf-8")
    write_master_cues_xlsx(cues)
