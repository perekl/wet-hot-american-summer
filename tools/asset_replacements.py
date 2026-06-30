#!/usr/bin/env python3
"""Export and apply sound-asset replacement worksheet.

Export a fillable spreadsheet of every unique sound file. After you add
``New Filename`` and/or ``New Download URL`` columns, run ``apply`` to
remap paths in assets.json, cues.json, asset_library.py, and master_sound_assets.xlsx.

Usage:
  python tools/asset_replacements.py export
  python tools/asset_replacements.py apply
  python tools/asset_replacements.py apply --download   # fetch New Download URLs
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
WORKSHEET = DATA / "asset_replacement_worksheet.xlsx"
ASSETS_JSON = DATA / "assets.json"
CUES_JSON = DATA / "cues.json"
ASSET_LIBRARY = ROOT / "tools" / "asset_library.py"

sys.path.insert(0, str(ROOT / "tools"))
sys.path.insert(0, str(ROOT / "app"))

from cue_script_index import cue_type  # noqa: E402
from generate_production_kit import (  # noqa: E402
    ASSET_COLS,
    _key_by_asset_id,
    write_xlsx,
)
from asset_library import CANONICAL_ASSETS  # noqa: E402

USER_AGENT = "WHAS-Production-Kit/1.0 (table-read; educational)"

COLUMNS = [
    ("Asset ID", 12),
    ("Name", 28),
    ("Role", 14),
    ("Category", 12),
    ("Playback Mode", 14),
    ("Cue Count", 10),
    ("Example Cues", 36),
    ("Current Filename", 40),
    ("Current Download URL", 40),
    ("New Filename", 40),
    ("New Download URL", 40),
    ("Your Notes", 30),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")


def load_assets() -> list[dict]:
    return json.loads(ASSETS_JSON.read_text(encoding="utf-8"))


def load_cues() -> list[dict]:
    return json.loads(CUES_JSON.read_text(encoding="utf-8"))


def _role_for_asset(asset_id: str, cues: list[dict]) -> str:
    bg = fg = False
    for cue in cues:
        if cue.get("asset_id") != asset_id:
            continue
        if cue_type(cue) == "BACKGROUND":
            bg = True
        else:
            fg = True
    if bg and fg:
        return "Both"
    if bg:
        return "Background"
    return "Foreground"


def _example_cues(asset_id: str, cues: list[dict], limit: int = 3) -> str:
    names = []
    for cue in cues:
        if cue.get("asset_id") == asset_id:
            names.append(f"{cue['id']}: {cue.get('name', '')[:40]}")
        if len(names) >= limit:
            break
    extra = sum(1 for c in cues if c.get("asset_id") == asset_id) - len(names)
    text = " | ".join(names)
    if extra > 0:
        text += f" | (+{extra} more)"
    return text


def _style_header_row(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _highlight_input_columns(ws, header_row: list[str]):
    input_cols = {"New Filename", "New Download URL", "Your Notes"}
    for idx, title in enumerate(header_row, start=1):
        if title in input_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).fill = INPUT_FILL


def export_worksheet(path: Path = WORKSHEET) -> Path:
    assets = load_assets()
    cues = load_cues()

    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Replacements"

    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for asset in assets:
        aid = asset["id"]
        ws.append([
            aid,
            asset.get("name", ""),
            _role_for_asset(aid, cues),
            asset.get("category", ""),
            asset.get("playback_mode", ""),
            len(asset.get("used_by", [])),
            _example_cues(aid, cues),
            asset.get("filename", ""),
            asset.get("download_url", ""),
            "",
            "",
            "",
        ])

    for idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    _highlight_input_columns(ws, headers)

    inst = wb.create_sheet("Instructions")
    inst["A1"] = "How to use this worksheet"
    inst["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "There are 77 unique sound files (76 real audio + 1 silence placeholder).",
        "239 cues reference these files — change the asset once and every cue updates.",
        "",
        "Fill in YELLOW columns on the Asset Replacements sheet:",
        "",
        "  New Filename — path relative to project root, e.g. assets/sfx/my_banana.wav",
        "               Leave blank to keep Current Filename.",
        "",
        "  New Download URL — optional direct link (mp3/wav). Used when you run:",
        "               python tools/asset_replacements.py apply --download",
        "",
        "  Your Notes — optional notes for yourself.",
        "",
        "You can fill only the rows you want to change; blank rows are skipped.",
        "",
        "After saving this file, run:",
        "  python tools/asset_replacements.py apply",
        "  python tools/asset_replacements.py apply --download   # also fetch URLs",
        "",
        "Or drop audio files directly onto Current Filename paths — no spreadsheet needed.",
    ]
    for i, line in enumerate(lines, start=2):
        inst.cell(row=i, column=1, value=line)
    inst.column_dimensions["A"].width = 90

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path} ({len(assets)} assets)")
    return path


def _download_file(url: str, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=120) as resp:
        dest.write_bytes(resp.read())


def _update_asset_library_filename(asset_id: str, new_filename: str) -> bool:
    key = _key_by_asset_id(asset_id)
    if key not in CANONICAL_ASSETS:
        return False
    old = CANONICAL_ASSETS[key]["filename"]
    if old == new_filename:
        return False
    CANONICAL_ASSETS[key]["filename"] = new_filename
    text = ASSET_LIBRARY.read_text(encoding="utf-8")
    pattern = (
        rf'("{re.escape(key)}":\s*\{{.*?"filename":\s*)"{re.escape(old)}"'
    )
    replacement = rf'\1"{new_filename}"'
    new_text, count = re.subn(pattern, replacement, text, count=1, flags=re.DOTALL)
    if count != 1:
        raise RuntimeError(
            f"Could not update asset_library.py for {asset_id} ({key}): "
            f"expected 1 replacement, got {count}"
        )
    ASSET_LIBRARY.write_text(new_text, encoding="utf-8")
    return True


def _assets_to_xlsx_rows(assets: list[dict]) -> list[dict]:
    rows = []
    for a in assets:
        used_by = a.get("used_by", [])
        if isinstance(used_by, list):
            used_by = ", ".join(used_by)
        rows.append({
            "Asset ID": a["id"],
            "Name": a.get("name", ""),
            "Category": a.get("category", ""),
            "Playback Mode": a.get("playback_mode", ""),
            "Suggested Volume": a.get("suggested_volume", ""),
            "Royalty Free?": a.get("royalty_free", ""),
            "Source": a.get("source", ""),
            "Filename": a.get("filename", ""),
            "Duration": a.get("duration", ""),
            "Used By Cue IDs": used_by,
            "Status": a.get("status", ""),
            "Download URL": a.get("download_url", ""),
            "Notes": a.get("notes", ""),
        })
    return rows


def apply_worksheet(path: Path = WORKSHEET, *, download: bool = False) -> None:
    if not path.is_file():
        raise SystemExit(f"Worksheet not found: {path}\nRun: python tools/asset_replacements.py export")

    wb = load_workbook(path, data_only=True)
    if "Asset Replacements" not in wb.sheetnames:
        raise SystemExit("Sheet 'Asset Replacements' not found in workbook")
    ws = wb["Asset Replacements"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {name: idx + 1 for idx, name in enumerate(headers) if name}

    required = {"Asset ID", "Current Filename", "New Filename", "New Download URL"}
    missing = required - set(col)
    if missing:
        raise SystemExit(f"Worksheet missing columns: {sorted(missing)}")

    asset_list = load_assets()
    assets = {a["id"]: a for a in asset_list}
    cues = load_cues()
    changes: list[dict] = []

    for row in range(2, ws.max_row + 1):
        asset_id = ws.cell(row, col["Asset ID"]).value
        if not asset_id:
            continue
        asset_id = str(asset_id).strip()
        if asset_id not in assets:
            print(f"  skip row {row}: unknown asset {asset_id}")
            continue

        new_filename = (ws.cell(row, col["New Filename"]).value or "").strip()
        new_url = (ws.cell(row, col["New Download URL"]).value or "").strip()
        if not new_filename and not new_url:
            continue

        current = assets[asset_id].get("filename", "")
        target = new_filename or current
        if not target:
            print(f"  skip row {row} ({asset_id}): no target filename")
            continue

        if download and new_url:
            dest = ROOT / target
            print(f"  downloading {asset_id} -> {target}")
            _download_file(new_url, dest)

        changes.append({
            "asset_id": asset_id,
            "filename": target,
            "download_url": new_url or assets[asset_id].get("download_url", ""),
        })

    if not changes:
        print("No changes found (fill New Filename and/or New Download URL).")
        return

    changed_ids = {c["asset_id"] for c in changes}
    change_by_id = {c["asset_id"]: c for c in changes}

    for asset in asset_list:
        if asset["id"] in change_by_id:
            ch = change_by_id[asset["id"]]
            asset["filename"] = ch["filename"]
            if ch["download_url"]:
                asset["download_url"] = ch["download_url"]
            dest = ROOT / asset["filename"]
            asset["status"] = "Sourced" if dest.is_file() else "Needed"

    cue_updates = 0
    for cue in cues:
        if cue.get("asset_id") in change_by_id:
            cue["asset_filename"] = change_by_id[cue["asset_id"]]["filename"]
            cue_updates += 1

    library_updates = 0
    for ch in changes:
        if _update_asset_library_filename(ch["asset_id"], ch["filename"]):
            library_updates += 1

    ASSETS_JSON.write_text(json.dumps(asset_list, indent=2), encoding="utf-8")
    CUES_JSON.write_text(json.dumps(cues, indent=2), encoding="utf-8")

    write_xlsx(
        DATA / "master_sound_assets.xlsx",
        "Sound Assets",
        ASSET_COLS,
        _assets_to_xlsx_rows(asset_list),
    )

    print(f"Applied {len(changes)} asset replacement(s):")
    for ch in changes:
        dest = ROOT / ch["filename"]
        on_disk = "yes" if dest.is_file() else "MISSING"
        print(f"  {ch['asset_id']} -> {ch['filename']}  (file on disk: {on_disk})")
    print(f"Updated {cue_updates} cue rows, {library_updates} asset_library.py entries, master_sound_assets.xlsx")


def main():
    parser = argparse.ArgumentParser(description="Export/apply asset replacement worksheet")
    sub = parser.add_subparsers(dest="command", required=True)
    exp = sub.add_parser("export", help="Write data/asset_replacement_worksheet.xlsx")
    exp.add_argument("--output", type=Path, default=WORKSHEET)
    app = sub.add_parser("apply", help="Apply filled worksheet to JSON + asset_library.py")
    app.add_argument("--worksheet", type=Path, default=WORKSHEET)
    app.add_argument(
        "--download", action="store_true",
        help="Download files from New Download URL into New Filename (or current path)",
    )
    args = parser.parse_args()
    if args.command == "export":
        export_worksheet(args.output)
    else:
        apply_worksheet(args.worksheet, download=args.download)


if __name__ == "__main__":
    main()
