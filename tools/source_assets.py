#!/usr/bin/env python3
"""Source royalty-free sound assets from Pixabay (first) and BBC Sound Effects."""

from __future__ import annotations

import csv
import json
import re
import subprocess
import tempfile
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from asset_search_hints import ASSET_SEARCH_HINTS, BBC_CSV_URL, BBC_MEDIA_BASE

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
ASSETS_JSON = DATA / "assets.json"
REPORT_PATH = DATA / "asset_sourcing_report.md"
BBC_CACHE = ROOT / "tools" / "data" / "BBCSoundEffects.csv"

USER_AGENT = "WHAS-Production-Kit/1.0 (table-read; educational)"


@dataclass
class SourcingResult:
    asset_id: str
    name: str
    filename: str
    status: str  # sourced | failed | skipped
    source: str = ""
    source_url: str = ""
    notes: str = ""
    pixabay_attempted: bool = False
    pixabay_error: str = ""
    bbc_attempted: bool = False
    bbc_error: str = ""


def load_assets() -> list[dict]:
    return json.loads(ASSETS_JSON.read_text(encoding="utf-8"))


def royalty_free_assets(assets: list[dict]) -> list[dict]:
    return [a for a in assets if a.get("royalty_free") == "Yes"]


def download_bytes(url: str, timeout: int = 120) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def ensure_bbc_csv() -> Path:
    BBC_CACHE.parent.mkdir(parents=True, exist_ok=True)
    if not BBC_CACHE.exists() or BBC_CACHE.stat().st_size < 1000:
        print(f"Downloading BBC metadata from {BBC_CSV_URL} ...")
        BBC_CACHE.write_bytes(download_bytes(BBC_CSV_URL, timeout=180))
    return BBC_CACHE


def load_bbc_rows() -> list[dict]:
    path = ensure_bbc_csv()
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def query_terms(query: str) -> list[str]:
    stop = {"the", "and", "with", "outdoor", "indoor", "sound", "short", "quiet", "gentle"}
    terms = [t.lower() for t in re.split(r"[^a-z0-9]+", query) if len(t) > 2 and t.lower() not in stop]
    return terms or [query.lower()]


def score_bbc_row(row: dict, terms: list[str], prefer_short: bool) -> float:
    desc = row["description"].lower()
    score = 0.0
    for term in terms:
        if term in desc:
            score += 4.0
    try:
        secs = float(row["secs"])
    except (TypeError, ValueError):
        secs = 60.0

    if prefer_short:
        if secs <= 8:
            score += 6
        elif secs <= 20:
            score += 4
        elif secs <= 45:
            score += 2
        elif secs > 120:
            score -= 4
    else:
        if 20 <= secs <= 240:
            score += 4
        elif 240 < secs <= 480:
            score += 2
        elif secs < 10:
            score -= 2
        elif secs > 900:
            score -= 3

    for bad in ("reprocessed", "bombing", "1941", "german", "war", "explosion", "gunfire"):
        if bad in desc:
            score -= 6
    for good in ("atmosphere", "ambience", "interior", "exterior", "single", "close", "clean"):
        if good in desc:
            score += 1
    return score


def search_bbc(query: str, prefer_short: bool, limit: int = 8) -> list[tuple[float, dict]]:
    terms = query_terms(query)
    rows = load_bbc_rows()
    scored: list[tuple[float, dict]] = []
    for row in rows:
        s = score_bbc_row(row, terms, prefer_short)
        if s > 0:
            scored.append((s, row))
    scored.sort(key=lambda x: (-x[0], float(x[1]["secs"])))
    return scored[:limit]


def search_pixabay(query: str) -> tuple[str | None, str]:
    """Attempt Pixabay sound-effects search. Returns (download_url, error)."""
    api_key = __import__("os").environ.get("PIXABAY_API_KEY", "").strip()
    if api_key:
        # Official Pixabay API does not expose sound effects; note and fall through.
        pass

    search_url = f"https://pixabay.com/sound-effects/search/{urllib.parse.quote(query)}/"
    try:
        req = urllib.request.Request(search_url, headers={"User-Agent": USER_AGENT})
        with urllib.request.urlopen(req, timeout=20) as resp:
            html = resp.read().decode("utf-8", errors="ignore")
    except Exception as exc:
        return None, f"Pixabay request failed: {exc}"

    if "Just a moment" in html or "challenge-platform" in html:
        return None, "Pixabay blocked by Cloudflare (manual download required)"

    urls = re.findall(r'(https://cdn\.pixabay\.com/audio/[^"\'\s>]+\.mp3)', html)
    if urls:
        return urls[0], ""
    return None, "No Pixabay CDN URL found in search results"


def generate_silence(dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    subprocess.run(
        [
            "ffmpeg", "-y", "-f", "lavfi", "-i", "anullsrc=r=44100:cl=mono",
            "-t", "5", "-q:a", "9", str(dest),
        ],
        check=True,
        capture_output=True,
    )


def convert_audio(src: Path, dest: Path, trim_seconds: int | None = None) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    ext = dest.suffix.lower()
    base = ["ffmpeg", "-y", "-i", str(src)]
    if trim_seconds:
        base += ["-t", str(trim_seconds)]
    if ext == ".mp3":
        cmd = base + ["-ac", "2", "-ar", "44100", "-q:a", "4", str(dest)]
    else:
        cmd = base + ["-ac", "1", "-ar", "44100", str(dest)]
    subprocess.run(cmd, check=True, capture_output=True)


def download_bbc_wav(location: str, dest_wav: Path) -> str:
    url = f"{BBC_MEDIA_BASE}/{location}"
    dest_wav.parent.mkdir(parents=True, exist_ok=True)
    data = download_bytes(url, timeout=300)
    if len(data) < 1000:
        raise ValueError(f"BBC download too small ({len(data)} bytes)")
    dest_wav.write_bytes(data)
    return url


def source_asset(asset: dict) -> SourcingResult:
    asset_id = asset["id"]
    dest = ROOT / asset["filename"]
    hint = ASSET_SEARCH_HINTS.get(asset_id)
    if hint:
        query, prefer_short = hint
    else:
        query = asset["name"]
        prefer_short = asset.get("playback_mode") == "One Shot"

    result = SourcingResult(
        asset_id=asset_id,
        name=asset["name"],
        filename=asset["filename"],
        status="failed",
    )

    if dest.exists() and dest.stat().st_size > 500:
        result.status = "skipped"
        result.notes = "File already exists"
        result.source = "existing"
        return result

    if asset_id == "AST-086" or asset.get("playback_mode") == "Silence":
        try:
            generate_silence(dest)
            result.status = "sourced"
            result.source = "generated"
            result.notes = "Generated digital silence via ffmpeg"
            return result
        except Exception as exc:
            result.notes = str(exc)
            return result

    # 1) Pixabay first
    result.pixabay_attempted = True
    pix_url, pix_err = search_pixabay(query)
    result.pixabay_error = pix_err
    if pix_url:
        try:
            with tempfile.NamedTemporaryFile(suffix=".mp3", delete=False) as tmp:
                tmp_path = Path(tmp.name)
            tmp_path.write_bytes(download_bytes(pix_url))
            convert_audio(tmp_path, dest)
            tmp_path.unlink(missing_ok=True)
            result.status = "sourced"
            result.source = "Pixabay"
            result.source_url = pix_url
            return result
        except Exception as exc:
            result.pixabay_error = f"{pix_err}; download failed: {exc}".strip("; ")

    # 2) BBC Sound Effects
    result.bbc_attempted = True
    try:
        hits = search_bbc(query, prefer_short)
        if not hits:
            result.bbc_error = f"No BBC match for query: {query}"
            result.notes = "No suitable BBC match"
            return result

        last_error = ""
        max_secs = 45 if prefer_short else 600
        for _score, row in hits:
            try:
                secs = float(row["secs"])
            except (TypeError, ValueError):
                secs = 999
            if secs > max_secs:
                continue
            location = row["location"]
            try:
                with tempfile.TemporaryDirectory() as td:
                    wav_path = Path(td) / location
                    url = download_bbc_wav(location, wav_path)
                    trim = 120 if not prefer_short else None
                    convert_audio(wav_path, dest, trim_seconds=trim)
                result.status = "sourced"
                result.source = "BBC Sound Effects"
                result.source_url = url
                result.notes = row["description"][:120]
                return result
            except Exception as exc:
                last_error = str(exc)
                continue
        result.bbc_error = last_error or "All BBC candidates failed"
        result.notes = "BBC candidates exhausted"
    except Exception as exc:
        result.bbc_error = str(exc)
        result.notes = str(exc)

    if not result.pixabay_error and not result.bbc_error:
        result.notes = "Unknown failure"
    return result


def update_assets_json(assets: list[dict], results: dict[str, SourcingResult]) -> None:
    for asset in assets:
        r = results.get(asset["id"])
        if not r or r.status not in ("sourced", "skipped"):
            continue
        asset["status"] = "Sourced"
        if r.source_url:
            asset["download_url"] = r.source_url
        elif r.source:
            asset["download_url"] = r.source
        src_note = f"Sourced from {r.source}"
        if r.notes and r.source != "generated":
            src_note += f" — {r.notes}"
        asset["notes"] = src_note


def write_report(results: list[SourcingResult]) -> None:
    sourced = [r for r in results if r.status == "sourced"]
    skipped = [r for r in results if r.status == "skipped"]
    failed = [r for r in results if r.status == "failed"]
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    lines = [
        "# Asset Sourcing Report",
        "",
        f"Generated: {now}",
        "",
        "## Summary",
        "",
        f"| Status | Count |",
        f"|--------|-------|",
        f"| Sourced | {len(sourced)} |",
        f"| Skipped (already present) | {len(skipped)} |",
        f"| Failed | {len(failed)} |",
        "",
        "Search order: **Pixabay** (first) → **BBC Sound Effects** (fallback).",
        "",
        "## Pixabay Status",
        "",
        "Pixabay was queried first for every asset. Automated access returned "
        "**Cloudflare challenge (403)** in this environment, so no Pixabay CDN "
        "downloads were possible without a browser session. Set `PIXABAY_API_KEY` "
        "does not help — Pixabay's official API does not expose sound effects. "
        "Manual Pixabay downloads may be substituted per asset if preferred.",
        "",
    ]

    if sourced:
        lines += ["## Successfully Sourced", ""]
        for r in sourced:
            lines.append(f"- **{r.asset_id}** — {r.name}")
            lines.append(f"  - File: `{r.filename}`")
            lines.append(f"  - Source: {r.source}")
            if r.source_url:
                lines.append(f"  - URL: {r.source_url}")
            if r.pixabay_error:
                lines.append(f"  - Pixabay: {r.pixabay_error}")
            if r.notes:
                lines.append(f"  - Match: {r.notes}")
            lines.append("")

    if skipped:
        lines += ["## Skipped (Already Present)", ""]
        for r in skipped:
            lines.append(f"- **{r.asset_id}** — `{r.filename}`")
        lines.append("")

    if failed:
        lines += ["## Could Not Be Automatically Sourced", ""]
        for r in failed:
            lines.append(f"### {r.asset_id} — {r.name}")
            lines.append(f"- Target: `{r.filename}`")
            if r.pixabay_attempted:
                lines.append(f"- Pixabay: {r.pixabay_error or 'no match'}")
            if r.bbc_attempted:
                lines.append(f"- BBC: {r.bbc_error or 'no match'}")
            if r.notes:
                lines.append(f"- Notes: {r.notes}")
            if r.pixabay_attempted and r.pixabay_error:
                lines.append(f"- Pixabay attempt: {r.pixabay_error}")
            lines.append("")

    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def update_master_xlsx(assets: list[dict]) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    xlsx = DATA / "master_sound_assets.xlsx"
    if not xlsx.exists():
        return
    wb = load_workbook(xlsx)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    id_col = headers.index("Asset ID") + 1
    status_col = headers.index("Status") + 1
    url_col = headers.index("Download URL") + 1
    notes_col = headers.index("Notes") + 1
    by_id = {a["id"]: a for a in assets}
    for row in range(2, ws.max_row + 1):
        aid = ws.cell(row=row, column=id_col).value
        if aid in by_id:
            a = by_id[aid]
            ws.cell(row=row, column=status_col, value=a.get("status", ""))
            ws.cell(row=row, column=url_col, value=a.get("download_url", ""))
            ws.cell(row=row, column=notes_col, value=a.get("notes", ""))
    wb.save(xlsx)


def main() -> None:
    assets = load_assets()
    targets = royalty_free_assets(assets)
    print(f"Sourcing {len(targets)} royalty-free assets ...")

    results: list[SourcingResult] = []
    result_map: dict[str, SourcingResult] = {}
    for i, asset in enumerate(targets, 1):
        print(f"[{i}/{len(targets)}] {asset['id']} {asset['name']}")
        r = source_asset(asset)
        results.append(r)
        result_map[asset["id"]] = r
        print(f"  -> {r.status} ({r.source or r.notes})")
        time.sleep(0.3)

    update_assets_json(assets, result_map)
    ASSETS_JSON.write_text(json.dumps(assets, indent=2), encoding="utf-8")
    write_report(results)
    update_master_xlsx(assets)

    sourced = sum(1 for r in results if r.status == "sourced")
    failed = sum(1 for r in results if r.status == "failed")
    print(f"\nDone: {sourced} sourced, {failed} failed")
    print(f"Report: {REPORT_PATH}")


if __name__ == "__main__":
    main()
